from __future__ import annotations

import csv
import io
import json
import re
from collections.abc import Iterable, Sequence
from pathlib import Path

from openpyxl import load_workbook

from usbc_average_lookup.models import InputBowler

_LINE_PATTERN = re.compile(r"^\s*(?P<name>.+?)\s*\(\s*[*_]*(?P<id>\d{4}-\d+)[*_]*\s*\)\s*$")
_NAME_HEADERS = {"name", "bowler", "bowler name", "bowlername", "fullname", "full name"}
_FIRST_HEADERS = {"first", "first name", "firstname", "given name"}
_LAST_HEADERS = {"last", "last name", "lastname", "surname", "family name"}
_ID_HEADERS = {
    "id",
    "member id",
    "memberid",
    "member number",
    "membership id",
    "membershipid",
    "membership number",
    "usbc id",
    "usbcid",
    "usbc number",
    "usbc id number",
}


def parse_input_file(path: Path, sheet_name: str | None = None) -> list[InputBowler]:
    suffix = path.suffix.casefold()
    if suffix == ".xlsx":
        return _parse_excel(path, sheet_name)
    if suffix == ".json":
        return parse_input_json(path.read_text(encoding="utf-8-sig"))
    return parse_input_text(path.read_text(encoding="utf-8-sig"))


def workbook_sheet_names(path: Path) -> list[str]:
    """Return non-empty worksheet names in workbook order."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return [
            sheet.title
            for sheet in workbook.worksheets
            if any(
                any(_cell_text(value) for value in row) for row in sheet.iter_rows(values_only=True)
            )
        ]
    finally:
        workbook.close()


def parse_input_text(text: str) -> list[InputBowler]:
    """Parse ``Name (prefix-suffix)`` lines or common delimited rows."""

    cleaned = [line.strip() for line in text.splitlines() if line.strip()]
    if not cleaned:
        return []

    combined = "\n".join(cleaned)
    if not any(delimiter in combined for delimiter in (",", "\t", "|", ";")):
        line_matches = [_LINE_PATTERN.match(line) for line in cleaned]
        return [
            InputBowler(
                match.group("name").strip() if match else line,
                match.group("id") if match else "",
            )
            for line, match in zip(cleaned, line_matches, strict=True)
        ]

    delimiter = _detect_delimiter(combined)
    rows = list(csv.reader(io.StringIO(combined), delimiter=delimiter))
    return _rows_to_bowlers(rows)


def parse_input_json(text: str) -> list[InputBowler]:
    try:
        payload = json.loads(text)
    except json.JSONDecodeError as error:
        raise ValueError("The JSON file is not valid") from error

    if isinstance(payload, dict):
        records = payload.get("bowlers", payload.get("results"))
        if records is None:
            records = [payload]
    else:
        records = payload
    if not isinstance(records, list):
        raise ValueError("The JSON file must contain a list of bowlers")

    bowlers: list[InputBowler] = []
    for position, record in enumerate(records, start=1):
        if isinstance(record, str):
            parsed = parse_input_text(record)
            if len(parsed) != 1:
                raise ValueError(f"JSON bowler {position} is not understandable")
            bowlers.append(parsed[0])
            continue
        if not isinstance(record, dict):
            raise ValueError(f"JSON bowler {position} must be a name or object")
        normalized = {_normalize_header(str(key)): value for key, value in record.items()}
        name = _first_value(normalized, _NAME_HEADERS)
        if not name:
            first = _first_value(normalized, _FIRST_HEADERS)
            last = _first_value(normalized, _LAST_HEADERS)
            name = " ".join(part for part in (first, last) if part).strip()
        membership_id = _clean_membership_id(_first_value(normalized, _ID_HEADERS))
        if not name and not membership_id:
            raise ValueError(f"JSON bowler {position} is missing a name or USBC ID")
        bowlers.append(InputBowler(name, membership_id))
    return bowlers


def _parse_excel(path: Path, sheet_name: str | None) -> list[InputBowler]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Excel sheet {sheet_name!r} was not found")
            sheets = [workbook[sheet_name]]
        else:
            sheets = workbook.worksheets
        for sheet in sheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            if any(any(_cell_text(value) for value in row) for row in rows):
                return _rows_to_bowlers(rows)
    finally:
        workbook.close()
    return []


def _rows_to_bowlers(rows: Sequence[Sequence[object]]) -> list[InputBowler]:
    nonempty = [list(row) for row in rows if any(_cell_text(value) for value in row)]
    if not nonempty:
        return []

    header = [_normalize_header(_cell_text(cell)) for cell in nonempty[0]]
    name_index = _find_header(header, _NAME_HEADERS)
    first_index = _find_header(header, _FIRST_HEADERS)
    last_index = _find_header(header, _LAST_HEADERS)
    id_index = _find_header(header, _ID_HEADERS)
    has_header = any(index is not None for index in (name_index, first_index, last_index, id_index))
    data_rows = nonempty[1:] if has_header else nonempty

    bowlers: list[InputBowler] = []
    for row_number, row in enumerate(data_rows, start=2 if has_header else 1):
        if name_index is not None:
            name = _at(row, name_index)
        elif first_index is not None or last_index is not None:
            name = " ".join(
                part for part in (_at(row, first_index), _at(row, last_index)) if part
            ).strip()
        elif not has_header:
            name = _at(row, 0)
        else:
            name = ""

        fallback_id_index = 1 if not has_header and len(row) > 1 else id_index
        membership_id = _clean_membership_id(_at(row, fallback_id_index))
        if not name and not membership_id:
            raise ValueError(f"Row {row_number} is missing a bowler name or USBC ID")
        bowlers.append(InputBowler(name, membership_id))
    return bowlers


def _detect_delimiter(text: str) -> str:
    try:
        return csv.Sniffer().sniff(text, delimiters=",\t|;").delimiter
    except csv.Error:
        return "\t"


def _find_header(header: list[str], aliases: set[str]) -> int | None:
    return next((index for index, value in enumerate(header) if value in aliases), None)


def _first_value(values: dict[str, object], aliases: Iterable[str]) -> str:
    for alias in aliases:
        value = values.get(alias)
        if value is not None and _cell_text(value):
            return _cell_text(value)
    return ""


def _at(row: Sequence[object], index: int | None) -> str:
    return _cell_text(row[index]) if index is not None and index < len(row) else ""


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def _normalize_header(value: str) -> str:
    return re.sub(r"[\s_-]+", " ", value.strip().casefold())


def _clean_membership_id(value: object) -> str:
    return _cell_text(value).strip("() ").strip("*_").strip()
