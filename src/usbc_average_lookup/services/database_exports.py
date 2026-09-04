"""Export stored identities and explicitly selected averages; never call the network."""

from __future__ import annotations

import csv
import json
import os
import re
import tempfile
from collections.abc import Iterable, Mapping
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from usbc_average_lookup.database import BowlerDatabase

BTM_HEADERS = (
    "First Name",
    "Middle Initial",
    "Last Name",
    "Gender",
    "Book Average",
    "Book Games",
    "USBC ID Number",
)
GENERIC_HEADERS = (
    *BTM_HEADERS,
    "Year",
    "Average Type",
    "Hand",
    "Association",
    "Association State",
    "Active",
    "Last Refreshed",
    "Status",
    "Notes",
    "Average Source",
    "League",
)


@dataclass(frozen=True)
class AverageRule:
    mode: str = "Latest"
    year: str = ""
    kind: str = "Standard"
    minimum_games: int = 1
    hand: str = "Any"
    source: str = "Composite"

    def __post_init__(self) -> None:
        if self.source not in {"Composite", "League", "Adjusted league"}:
            raise ValueError("Choose a valid average source")
        if self.mode not in {"Latest", "Highest", "Specific year", "Manual"}:
            raise ValueError("Choose a valid average rule")
        if self.mode == "Specific year" and not self.year:
            raise ValueError("Choose a stored year")
        if self.kind not in {"Standard", "Sport", "Challenge", "Any"}:
            raise ValueError("Choose a valid average type")
        if self.minimum_games < 0:
            raise ValueError("Minimum games must be zero or greater")


def average_type(row: Mapping) -> str:
    types = [name for key, name in (("sport", "Sport"), ("challenge", "Challenge")) if row[key]]
    return " + ".join(types) or "Standard"


def year_key(value: str) -> tuple[int, ...]:
    return tuple(int(part) for part in re.findall(r"\d+", value)) or (0,)


def select_average(
    averages: Iterable[dict], rule: AverageRule, manual_id: int | None = None
) -> dict | None:
    rows = list(averages)
    if rule.mode == "Manual":
        return next((row for row in rows if row["id"] == manual_id), None)
    rows = [
        row
        for row in rows
        if row["games"] >= rule.minimum_games
        and (rule.kind == "Any" or average_type(row) == rule.kind)
        and (rule.hand == "Any" or row["hand"] == rule.hand)
        and (rule.mode != "Specific year" or row["year"] == rule.year)
    ]
    if not rows:
        return None

    def ordering(row):
        latest = (year_key(row["year"]), row["average"], row["games"], row["hand"], row["id"])
        return (row["average"], *latest) if rule.mode == "Highest" else latest

    return max(rows, key=ordering)


def export_preview(
    database: BowlerDatabase,
    bowler_ids: Iterable[int],
    rule: AverageRule,
    manual: Mapping[int, int] | None = None,
) -> list[dict]:
    preview = []
    for bowler_id in dict.fromkeys(bowler_ids):
        bowler = database.get(bowler_id)
        average = select_average(
            stored_averages(database, bowler_id, rule.source), rule, (manual or {}).get(bowler_id)
        )
        preview.append({"bowler": bowler, "average": average})
    return preview


def stored_averages(database: BowlerDatabase, bowler_id: int, source: str) -> list[dict]:
    rows = (
        database.averages(bowler_id)
        if source == "Composite"
        else database.league_averages(bowler_id)
    )
    if source == "Adjusted league":
        rows = [
            {**row, "average": row["adjusted_average"]}
            for row in rows
            if row["adjusted_average"] > 0
        ]
    return [{**row, "source": source} for row in rows]


def _safe_cell(value: object) -> object:
    # Spreadsheet programs should treat user/source strings as literal text.
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def export_database(
    path: Path, preview: list[dict], *, format: str = "BTM26+", missing: str = "Error"
) -> int:
    if format not in {"BTM26+", "CSV", "XLSX", "JSON"}:
        raise ValueError("Choose BTM26+, CSV, XLSX, or JSON")
    if missing not in {"Error", "Blank", "Skip"}:
        raise ValueError("Choose how to handle missing averages")
    if missing == "Error" and any(item["average"] is None for item in preview):
        raise ValueError("Some bowlers have no matching average. Choose Blank or Skip to continue.")
    selected = [item for item in preview if item["average"] is not None or missing != "Skip"]
    if not selected:
        raise ValueError("No bowlers match this export")
    headers = BTM_HEADERS if format == "BTM26+" else GENERIC_HEADERS
    rows = []
    for item in selected:
        b, a = item["bowler"], item["average"]
        row = [
            b["first_name"],
            b["middle_initial"][:1],
            b["last_name"],
            b["gender"],
            a["average"] if a else "",
            a["games"] if a else "",
            b["membership_id"] or "",
        ]
        if format != "BTM26+":
            row.extend(
                [
                    a["year"] if a else "",
                    average_type(a) if a else "",
                    a["hand"] if a else "",
                    b["association"],
                    b["association_state"],
                    "" if b["active"] is None else "Yes" if b["active"] else "No",
                    b["refreshed_at"] or "",
                    b["status"],
                    b["note"],
                    a.get("source", "Composite") if a else "",
                    a.get("league_name", "") if a else "",
                ]
            )
        rows.append([_safe_cell(value) for value in row])
    # Atomic replacement: a failed save never destroys an existing export.
    descriptor, temporary = tempfile.mkstemp(dir=path.parent, suffix=path.suffix)
    os.close(descriptor)
    target = Path(temporary)
    try:
        if format == "XLSX":
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Bowlers"
            sheet.append(headers)
            for row in rows:
                sheet.append(row)
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="153252")
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cells in sheet.columns:
                sheet.column_dimensions[cells[0].column_letter].width = min(
                    42, max(len(str(cell.value or "")) for cell in cells) + 2
                )
            workbook.save(target)
            workbook.close()
        elif format == "JSON":
            document = {
                "schemaVersion": 3,
                "bowlers": [
                    {
                        "name": item["bowler"]["display_name"],
                        "membershipId": item["bowler"]["membership_id"],
                        "member": item["bowler"],
                        "selectedAverage": item["average"],
                    }
                    for item in selected
                ],
            }
            target.write_text(json.dumps(document, ensure_ascii=False, indent=2), encoding="utf-8")
        else:
            with target.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(headers)
                writer.writerows(rows)
        target.replace(path)
    finally:
        target.unlink(missing_ok=True)
    return len(rows)
