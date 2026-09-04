from __future__ import annotations

import csv
import json
from collections import Counter
from collections.abc import Iterable
from datetime import UTC, datetime
from enum import StrEnum
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from usbc_average_lookup.models import LookupResult, LookupStatus

_HEADERS = (
    "Name",
    "Member ID",
    "Average",
    "Year",
    "Games",
    "Status",
    "Notes",
    "Active",
    "Association",
)


class ExportSubset(StrEnum):
    FULL = "Full roster"
    READY = "Active / ready roster"
    INACTIVE = "Inactive roster"
    NEEDS_ATTENTION = "Needs-attention roster"


def select_results(
    results: Iterable[LookupResult], subset: ExportSubset = ExportSubset.FULL
) -> list[LookupResult]:
    result_list = list(results)
    if subset is ExportSubset.FULL:
        return result_list
    if subset is ExportSubset.READY:
        return [
            result
            for result in result_list
            if result.status is LookupStatus.FOUND
            and (result.member is None or result.member.active)
        ]
    if subset is ExportSubset.INACTIVE:
        return [
            result
            for result in result_list
            if result.status is LookupStatus.INACTIVE_MEMBER
            or (result.member is not None and not result.member.active)
        ]
    return [result for result in result_list if result.needs_attention]


def export_results(
    path: Path,
    results: Iterable[LookupResult],
    subset: ExportSubset = ExportSubset.FULL,
) -> int:
    selected = select_results(results, subset)
    suffix = path.suffix.casefold()
    if suffix == ".json":
        export_json(path, selected, subset)
    elif suffix == ".xlsx":
        export_excel(path, selected)
    elif suffix == ".csv":
        export_delimited(path, selected, ",")
    elif suffix in {".tsv", ".txt"}:
        export_delimited(path, selected, "\t")
    else:
        raise ValueError("Choose JSON, CSV, TSV, TXT, or Excel output")
    return len(selected)


def export_delimited(path: Path, results: Iterable[LookupResult], delimiter: str) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle, delimiter=delimiter)
        writer.writerow(_HEADERS)
        writer.writerows(_row(result) for result in results)


def export_excel(path: Path, results: Iterable[LookupResult]) -> None:
    result_list = list(results)
    workbook = Workbook()
    results_sheet = workbook.active
    results_sheet.title = "Results"
    _write_sheet(results_sheet, result_list)
    issues = [result for result in result_list if result.needs_attention]
    if issues:
        _write_sheet(workbook.create_sheet("Needs Attention"), issues)
    workbook.save(path)


def export_found(path: Path, results: Iterable[LookupResult]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Name", "Average"])
        writer.writerows(
            (result.input_name, result.average)
            for result in results
            if result.status is LookupStatus.FOUND
        )


def export_issues(path: Path, results: Iterable[LookupResult]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Name", "Status", "Notes"])
        writer.writerows(
            (result.input_name, result.status.value, result.note)
            for result in results
            if result.status is not LookupStatus.FOUND
        )


def export_json(
    path: Path,
    results: Iterable[LookupResult],
    subset: ExportSubset = ExportSubset.FULL,
) -> None:
    result_list = list(results)
    counts = Counter(result.status for result in result_list)
    document = {
        "schemaVersion": 2,
        "generatedAt": datetime.now(UTC).isoformat(),
        "rosterType": subset.value,
        "summary": {
            "processed": len(result_list),
            **{status.name.lower(): counts[status] for status in LookupStatus},
        },
        "bowlers": [
            {
                "name": result.input_name,
                "membershipId": result.membership_id or None,
                "average": result.average,
                "year": result.year or None,
                "games": result.games,
                "status": result.status.value,
                "notes": result.note or None,
                "active": result.member.active if result.member else None,
                "association": result.member.association if result.member else None,
                "associationState": (result.member.association_state if result.member else None),
            }
            for result in result_list
        ],
    }
    path.write_text(json.dumps(document, indent=2) + "\n", encoding="utf-8")


def _row(result: LookupResult) -> tuple[object, ...]:
    member = result.member
    return (
        result.input_name,
        result.membership_id,
        result.average if result.average is not None else "",
        result.year,
        result.games if result.games is not None else "",
        result.status.value,
        result.note,
        "Yes" if member and member.active else "No" if member else "",
        ", ".join(part for part in (member.association, member.association_state) if part)
        if member
        else "",
    )


def _write_sheet(sheet, results: list[LookupResult]) -> None:
    sheet.append(_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="153252")
    for result in results:
        sheet.append(_row(result))
        status_cell = sheet.cell(sheet.max_row, 6)
        if result.status is LookupStatus.FOUND:
            status_cell.fill = PatternFill("solid", fgColor="E4F3EA")
        elif result.status is LookupStatus.INACTIVE_MEMBER:
            status_cell.fill = PatternFill("solid", fgColor="F8EDD6")
        else:
            status_cell.fill = PatternFill("solid", fgColor="FAE8E6")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, header in enumerate(_HEADERS, start=1):
        values = [str(sheet.cell(row, index).value or "") for row in range(1, sheet.max_row + 1)]
        width = min(max(len(header), *(len(value) for value in values)) + 2, 42)
        sheet.column_dimensions[get_column_letter(index)].width = width
