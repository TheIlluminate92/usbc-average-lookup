from usbc_average_lookup.models import InputBowler
import json

from openpyxl import Workbook

from usbc_average_lookup.services.input_parser import (
    parse_input_file,
    parse_input_json,
    parse_input_text,
    workbook_sheet_names,
)


def test_parses_name_and_parenthesized_membership_id_lines() -> None:
    text = """Alex Bowler (1234-567890)
Jamie Bowler (**9876-54321**)
"""

    assert parse_input_text(text) == [
        InputBowler("Alex Bowler", "1234-567890"),
        InputBowler("Jamie Bowler", "9876-54321"),
    ]


def test_parses_csv_with_headers() -> None:
    text = "Name,Membership ID\nAlex Bowler,1234-567890\nJamie Bowler,\n"

    assert parse_input_text(text) == [
        InputBowler("Alex Bowler", "1234-567890"),
        InputBowler("Jamie Bowler", ""),
    ]


def test_parses_headerless_pipe_delimited_text() -> None:
    text = "Alex Bowler|1234-567890\nJamie Bowler|9876-54321\n"

    assert parse_input_text(text) == [
        InputBowler("Alex Bowler", "1234-567890"),
        InputBowler("Jamie Bowler", "9876-54321"),
    ]


def test_parses_separate_first_and_last_name_columns() -> None:
    text = "First Name,Last Name,USBC ID\nAlex,Bowler,1234-567890\n"

    assert parse_input_text(text) == [InputBowler("Alex Bowler", "1234-567890")]


def test_parses_json_list_and_export_shape() -> None:
    direct = json.dumps(
        [
            {"name": "Alex Bowler", "membershipId": "1234-567890"},
            {"firstName": "Jamie", "lastName": "Bowler"},
        ]
    )

    assert parse_input_json(direct) == [
        InputBowler("Alex Bowler", "1234-567890"),
        InputBowler("Jamie Bowler", ""),
    ]
    assert parse_input_json(json.dumps({"bowlers": json.loads(direct)})) == [
        InputBowler("Alex Bowler", "1234-567890"),
        InputBowler("Jamie Bowler", ""),
    ]


def test_parses_excel_and_lists_nonempty_sheets(tmp_path) -> None:
    path = tmp_path / "roster.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "League A"
    first.append(["Name", "Membership ID"])
    first.append(["Alex Bowler", "1234-567890"])
    workbook.create_sheet("Empty")
    second = workbook.create_sheet("League B")
    second.append(["First Name", "Last Name", "USBC ID"])
    second.append(["Jamie", "Bowler", "9876-54321"])
    workbook.save(path)

    assert workbook_sheet_names(path) == ["League A", "League B"]
    assert parse_input_file(path) == [InputBowler("Alex Bowler", "1234-567890")]
    assert parse_input_file(path, "League B") == [
        InputBowler("Jamie Bowler", "9876-54321")
    ]
