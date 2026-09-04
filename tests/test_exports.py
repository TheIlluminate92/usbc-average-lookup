import csv

from openpyxl import load_workbook

from usbc_average_lookup.models import LookupResult, LookupStatus, Member
from usbc_average_lookup.services.exports import (
    ExportSubset,
    export_found,
    export_issues,
    export_results,
    select_results,
)


def read_rows(path):
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.reader(handle))


def test_exports_found_and_issues_separately(tmp_path) -> None:
    results = [
        LookupResult("John Smith", LookupStatus.FOUND, average=187),
        LookupResult("Jane Doe", LookupStatus.NOT_FOUND, note="No matching member"),
    ]
    found_path = tmp_path / "results.csv"
    issues_path = tmp_path / "issues.csv"

    export_found(found_path, results)
    export_issues(issues_path, results)

    assert read_rows(found_path) == [["Name", "Average"], ["John Smith", "187"]]
    assert read_rows(issues_path) == [
        ["Name", "Status", "Notes"],
        ["Jane Doe", "Not found", "No matching member"],
    ]


def test_selects_full_ready_inactive_and_attention_rosters() -> None:
    inactive_member = Member("2", "1234", "2", "Former", "Bowler", False)
    results = [
        LookupResult("Ready", LookupStatus.FOUND, average=180),
        LookupResult(
            "Inactive",
            LookupStatus.INACTIVE_MEMBER,
            average=150,
            member=inactive_member,
            confirmed_inactive=True,
        ),
        LookupResult("Missing", LookupStatus.NOT_FOUND),
    ]

    assert len(select_results(results, ExportSubset.FULL)) == 3
    assert [item.input_name for item in select_results(results, ExportSubset.READY)] == ["Ready"]
    assert [item.input_name for item in select_results(results, ExportSubset.INACTIVE)] == [
        "Inactive"
    ]
    assert [item.input_name for item in select_results(results, ExportSubset.NEEDS_ATTENTION)] == [
        "Missing"
    ]


def test_exports_csv_tsv_and_excel_with_issue_sheet(tmp_path) -> None:
    results = [
        LookupResult("Ready Bowler", LookupStatus.FOUND, average=180, year="2025", games=60),
        LookupResult("Missing Bowler", LookupStatus.NOT_FOUND, note="Check spelling"),
    ]
    csv_path = tmp_path / "results.csv"
    tsv_path = tmp_path / "results.tsv"
    excel_path = tmp_path / "results.xlsx"

    assert export_results(csv_path, results) == 2
    assert export_results(tsv_path, results) == 2
    assert export_results(excel_path, results) == 2

    assert read_rows(csv_path)[1][:7] == [
        "Ready Bowler",
        "",
        "180",
        "2025",
        "60",
        "Found",
        "",
    ]
    with tsv_path.open(encoding="utf-8-sig") as handle:
        assert handle.readline().startswith("Name\tMember ID\tAverage")
    workbook = load_workbook(excel_path, read_only=True)
    try:
        assert workbook.sheetnames == ["Results", "Needs Attention"]
        assert workbook["Needs Attention"]["A2"].value == "Missing Bowler"
    finally:
        workbook.close()
